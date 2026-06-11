import streamlit as st
import openpyxl
import plotly.graph_objects as go

from database import criar_banco, get_conn

st.set_page_config(page_title="Banco de Questões", layout="wide")


# ===================================================
# LOGIN
# ===================================================

if "logado" not in st.session_state:
    st.session_state.logado = False

if not st.session_state.logado:

    st.title("🔒 Banco de Questões")

    senha = st.text_input("Senha", type="password")

    if st.button("Entrar"):
        if senha == st.secrets["APP_PASSWORD"]:
            st.session_state.logado = True
            st.rerun()
        else:
            st.error("Senha incorreta.")

    st.stop()


# ===================================================
# INICIALIZAÇÃO DO BANCO (só uma vez por sessão)
# ===================================================

if "db_iniciado" not in st.session_state:
    try:
        criar_banco()
        st.session_state["db_iniciado"] = True
    except Exception as e:
        st.error(f"Erro ao conectar no Supabase: {e}")
        st.stop()


# ===================================================
# CONFIGURAÇÃO DA PÁGINA
# ===================================================

st.title("📚 Banco de Questões Para Seu Futuro Cargo")

menu_opcoes = [
    "Cadastrar Questão",
    "Importar por Excel",
    "Gerenciar Questões",
    "Resolver Questões",
    "Revisar Questões",
    "Estatísticas",
    "Dashboard",
]

indice_menu = 0

if "menu_override" in st.session_state:
    try:
        indice_menu = menu_opcoes.index(st.session_state.pop("menu_override"))
    except ValueError:
        indice_menu = 0

menu = st.sidebar.selectbox("Menu", menu_opcoes, index=indice_menu)


# ===================================================
# FUNÇÕES AUXILIARES
# ===================================================

def get_todas_tags():
    """Retorna lista ordenada de todas as tags únicas do banco."""
    conn   = get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT tags FROM questoes")
    registros = cursor.fetchall()
    cursor.close()
    conn.close()

    lista = []
    for (tags_raw,) in registros:
        if tags_raw:
            for tag in tags_raw.splitlines():
                tag = tag.strip()
                if tag and tag not in lista:
                    lista.append(tag)
    lista.sort()
    return lista


def atualizar_estatisticas_tag(cursor, tags, novo_status):
    """Atualiza acertos/erros por tag no ciclo atual."""
    for tag in tags.splitlines():
        tag = tag.strip()
        if not tag:
            continue

        cursor.execute("""
            SELECT ciclo FROM estatisticas_tag
            WHERE tag = %s
            ORDER BY ciclo DESC
            LIMIT 1
        """, (tag,))

        row = cursor.fetchone()

        if row:
            ciclo_atual = row[0]
        else:
            ciclo_atual = 1
            cursor.execute("""
                INSERT INTO estatisticas_tag (tag, ciclo, acertos, erros)
                VALUES (%s, %s, 0, 0)
            """, (tag, ciclo_atual))

        if novo_status == "Acertada":
            cursor.execute("""
                UPDATE estatisticas_tag
                SET acertos = acertos + 1
                WHERE tag = %s AND ciclo = %s
            """, (tag, ciclo_atual))
        else:
            cursor.execute("""
                UPDATE estatisticas_tag
                SET erros = erros + 1
                WHERE tag = %s AND ciclo = %s
            """, (tag, ciclo_atual))


def nivel_dominio(pct):
    if pct <= 50:  return "🔴 Crítico"
    if pct <= 70:  return "🟠 Fraco"
    if pct <= 85:  return "🟡 Bom"
    if pct <= 95:  return "🟢 Muito Bom"
    return "🏆 Dominado"


def cor_dominio(pct):
    if pct <= 50:  return "#e74c3c"
    if pct <= 70:  return "#e67e22"
    if pct <= 85:  return "#f1c40f"
    if pct <= 95:  return "#2ecc71"
    return "#1abc9c"


def editor_com_negrito(label, value="", key=""):
    """
    Editor de texto com suporte a negrito usando marcação **texto**.
    Exibe uma barra de ferramentas com botão de negrito e um text_area.
    Retorna o texto com marcação **...**  que é salvo no banco.
    """
    st.markdown(f"**{label}**")

    col_btn, col_info = st.columns([1, 5])
    with col_btn:
        st.markdown(
            """
            <div style='margin-bottom:4px;'>
                <span style='background:#f0f0f0; border:1px solid #ccc;
                             border-radius:4px; padding:2px 10px;
                             font-weight:bold; font-size:14px;'>B</span>
                <span style='font-size:12px; color:#888; margin-left:6px;'>
                    Selecione o texto e envolva com **texto**
                </span>
            </div>
            """,
            unsafe_allow_html=True
        )

    texto = st.text_area(
        label,
        value=value,
        key=key,
        label_visibility="collapsed",
        help="Use **texto** para negrito. Ex: **importante**"
    )
    return texto


def renderizar_com_negrito(texto):
    """
    Renderiza texto com suporte a marcação **negrito** usando st.markdown.
    Preserva quebras de linha.
    """
    if not texto:
        return
    # Converte quebras de linha em markdown
    texto_md = texto.replace("\n", "  \n")
    st.markdown(texto_md)


# ===================================================
# CADASTRAR QUESTÃO
# ===================================================

if menu == "Cadastrar Questão":

    st.header("Nova Questão")

    # Inicializa contadores de reset
    if "cadastro_reset" not in st.session_state:
        st.session_state["cadastro_reset"] = 0

    reset_key = st.session_state["cadastro_reset"]

    materia    = st.text_input("Matéria",    key=f"cad_materia_{reset_key}")
    assunto    = st.text_input("Assunto",    key=f"cad_assunto_{reset_key}")
    banca      = st.text_input("Banca",      key=f"cad_banca_{reset_key}")
    cargo      = st.text_input("Cargo",      key=f"cad_cargo_{reset_key}")
    ano        = st.number_input("Ano", min_value=2000, max_value=2100, value=2025,
                                 key=f"cad_ano_{reset_key}")
    dificuldade = st.selectbox("Dificuldade",
        ["Muito Fácil", "Fácil", "Média", "Difícil", "Muito Difícil", "Pegadinha"],
        key=f"cad_dif_{reset_key}")
    tags       = st.text_area("Tags (uma por linha)", height=120,
                              key=f"cad_tags_{reset_key}")

    questao    = editor_com_negrito("Questão", key=f"cad_questao_{reset_key}")
    gabarito   = st.selectbox("Gabarito", ["Certo", "Errado"],
                              key=f"cad_gab_{reset_key}")
    comentario = editor_com_negrito("Comentário", key=f"cad_coment_{reset_key}")

    if st.button("Salvar"):
        if not questao.strip():
            st.warning("O campo Questão não pode estar vazio.")
        else:
            conn   = get_conn()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO questoes
                (materia, assunto, banca, cargo, ano, dificuldade,
                 tags, questao, gabarito, comentario, observacoes, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (materia, assunto, banca, cargo, ano, dificuldade,
                  tags, questao, gabarito, comentario, "", "Não respondida"))
            conn.commit()
            cursor.close()
            conn.close()
            st.success("✅ Questão salva com sucesso!")
            # Reseta o formulário incrementando a chave
            st.session_state["cadastro_reset"] += 1
            st.rerun()


# ===================================================
# IMPORTAR POR EXCEL
# ===================================================

elif menu == "Importar por Excel":

    st.header("📥 Importação em Lote por Excel")

    st.markdown("""
**Formato esperado (uma questão por linha):**

| materia | assunto | banca | cargo | ano | dificuldade | tags | questao | gabarito | comentario |
|---------|---------|-------|-------|-----|-------------|------|---------|----------|------------|

- **dificuldade**: Muito Fácil / Fácil / Média / Difícil / Muito Difícil / Pegadinha
- **gabarito**: Certo / Errado
- **tags**: separe com ponto e vírgula (`;`)
- A primeira linha deve ser o cabeçalho
- Use `**texto**` no campo questao/comentario para negrito
    """)

    arquivo = st.file_uploader("Selecione o arquivo Excel (.xlsx)", type=["xlsx"])

    if arquivo:
        try:
            wb = openpyxl.load_workbook(arquivo)
            ws = wb.active

            cabecalho = [
                str(c.value).strip().lower() if c.value else ""
                for c in ws[1]
            ]

            colunas_esperadas = [
                "materia","assunto","banca","cargo","ano",
                "dificuldade","tags","questao","gabarito","comentario"
            ]

            faltando = [c for c in colunas_esperadas if c not in cabecalho]

            if faltando:
                st.error(f"Colunas faltando: {', '.join(faltando)}")
            else:
                idx = {c: cabecalho.index(c) for c in colunas_esperadas}
                linhas = list(ws.iter_rows(min_row=2, values_only=True))
                linhas_validas = [l for l in linhas if any(c for c in l if c is not None)]

                st.info(f"{len(linhas_validas)} questões encontradas no arquivo.")

                if st.button("✅ Confirmar Importação"):
                    conn   = get_conn()
                    cursor = conn.cursor()
                    importadas = 0
                    erros_imp  = []

                    for i, linha in enumerate(linhas_validas, start=2):
                        try:
                            tags_raw = str(linha[idx["tags"]] or "").strip()
                            tags_fmt = "\n".join(
                                t.strip() for t in tags_raw.split(";") if t.strip()
                            )
                            cursor.execute("""
                                INSERT INTO questoes
                                (materia,assunto,banca,cargo,ano,dificuldade,
                                 tags,questao,gabarito,comentario,observacoes,status)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """, (
                                str(linha[idx["materia"]]     or "").strip(),
                                str(linha[idx["assunto"]]     or "").strip(),
                                str(linha[idx["banca"]]       or "").strip(),
                                str(linha[idx["cargo"]]       or "").strip(),
                                int(linha[idx["ano"]]         or 2025),
                                str(linha[idx["dificuldade"]] or "Média").strip(),
                                tags_fmt,
                                str(linha[idx["questao"]]     or "").strip(),
                                str(linha[idx["gabarito"]]    or "Certo").strip(),
                                str(linha[idx["comentario"]]  or "").strip(),
                                "", "Não respondida"
                            ))
                            importadas += 1
                        except Exception as e:
                            erros_imp.append(f"Linha {i}: {e}")

                    conn.commit()
                    cursor.close()
                    conn.close()

                    st.success(f"✅ {importadas} questões importadas!")
                    for erro in erros_imp:
                        st.warning(f"• {erro}")

        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")

    st.divider()
    st.subheader("📄 Baixar Modelo de Planilha")

    if st.button("Gerar modelo Excel"):
        wb_m = openpyxl.Workbook()
        ws_m = wb_m.active
        ws_m.title = "Questoes"
        ws_m.append(["materia","assunto","banca","cargo","ano",
                      "dificuldade","tags","questao","gabarito","comentario"])
        ws_m.append(["Direito Civil","Curatela","CESPE","Analista",2024,"Média",
                      "Curatela;Incapacidade",
                      "A curatela é medida protetiva de caráter excepcional.",
                      "Certo","Conforme art. 84 do Estatuto da Pessoa com Deficiência."])
        caminho = "/tmp/modelo_questoes.xlsx"
        wb_m.save(caminho)
        with open(caminho, "rb") as f:
            st.download_button("⬇️ Baixar modelo.xlsx", f,
                file_name="modelo_questoes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ===================================================
# GERENCIAR QUESTÕES
# ===================================================

elif menu == "Gerenciar Questões":

    st.header("⚙️ Gerenciar Questões")

    # --- FILTROS DE BUSCA ---
    with st.expander("🔍 Filtros de Busca", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            gq_materia = st.text_input("Matéria", key="gq_materia")
            gq_banca   = st.text_input("Banca",   key="gq_banca")
        with col2:
            gq_cargo   = st.text_input("Cargo",   key="gq_cargo")
            gq_ano     = st.text_input("Ano",     key="gq_ano")
        with col3:
            gq_dif     = st.selectbox("Dificuldade",
                ["","Muito Fácil","Fácil","Média","Difícil","Muito Difícil","Pegadinha"],
                key="gq_dif")
            gq_status  = st.selectbox("Status",
                ["","Não respondida","Acertada","Errada"], key="gq_status")

        gq_texto = st.text_input("Buscar no texto da questão", key="gq_texto")

        if st.button("🔍 Buscar"):
            sql, params = "SELECT id, materia, assunto, banca, cargo, ano, dificuldade, questao, gabarito, status, dominio FROM questoes WHERE 1=1", []
            if gq_materia: sql += " AND materia ILIKE %s"; params.append(f"%{gq_materia}%")
            if gq_banca:   sql += " AND banca ILIKE %s";   params.append(f"%{gq_banca}%")
            if gq_cargo:   sql += " AND cargo ILIKE %s";   params.append(f"%{gq_cargo}%")
            if gq_ano:     sql += " AND ano = %s";         params.append(gq_ano)
            if gq_dif:     sql += " AND dificuldade = %s"; params.append(gq_dif)
            if gq_status:  sql += " AND status = %s";      params.append(gq_status)
            if gq_texto:   sql += " AND questao ILIKE %s"; params.append(f"%{gq_texto}%")
            sql += " ORDER BY id DESC"

            conn   = get_conn()
            cursor = conn.cursor()
            cursor.execute(sql, params)
            st.session_state["gq_resultados"] = cursor.fetchall()
            cursor.close()
            conn.close()

    # --- LISTAGEM ---
    if "gq_resultados" in st.session_state:
        resultados = st.session_state["gq_resultados"]

        if not resultados:
            st.warning("Nenhuma questão encontrada.")
        else:
            st.success(f"{len(resultados)} questão(ões) encontrada(s).")

            # Se está editando uma questão específica
            if "gq_editando_id" in st.session_state:
                qid = st.session_state["gq_editando_id"]

                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM questoes WHERE id = %s", (qid,))
                q = cursor.fetchone()
                cursor.close()
                conn.close()

                if q:
                    st.divider()
                    st.subheader(f"✏️ Editando Questão #{q[0]}")

                    dif_opcoes = ["Muito Fácil","Fácil","Média","Difícil","Muito Difícil","Pegadinha"]
                    gab_opcoes = ["Certo","Errado"]

                    e_materia    = st.text_input("Matéria",  value=q[1] or "", key="ed_materia")
                    e_assunto    = st.text_input("Assunto",  value=q[2] or "", key="ed_assunto")
                    e_banca      = st.text_input("Banca",    value=q[3] or "", key="ed_banca")
                    e_cargo      = st.text_input("Cargo",    value=q[4] or "", key="ed_cargo")
                    e_ano        = st.number_input("Ano", min_value=2000, max_value=2100,
                                                   value=int(q[5] or 2025), key="ed_ano")
                    e_dif        = st.selectbox("Dificuldade", dif_opcoes,
                                                index=dif_opcoes.index(q[6]) if q[6] in dif_opcoes else 2,
                                                key="ed_dif")
                    e_tags       = st.text_area("Tags (uma por linha)", value=q[7] or "",
                                                height=100, key="ed_tags")

                    e_questao    = editor_com_negrito("Questão",    value=q[8]  or "", key="ed_questao")
                    e_gab        = st.selectbox("Gabarito", gab_opcoes,
                                                index=gab_opcoes.index(q[9]) if q[9] in gab_opcoes else 0,
                                                key="ed_gab")
                    e_comentario = editor_com_negrito("Comentário", value=q[10] or "", key="ed_coment")
                    e_obs        = st.text_area("Observações", value=q[11] or "",
                                               height=80, key="ed_obs")

                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("💾 Salvar Alterações"):
                            conn   = get_conn()
                            cursor = conn.cursor()
                            cursor.execute("""
                                UPDATE questoes SET
                                    materia=%s, assunto=%s, banca=%s, cargo=%s,
                                    ano=%s, dificuldade=%s, tags=%s, questao=%s,
                                    gabarito=%s, comentario=%s, observacoes=%s
                                WHERE id=%s
                            """, (e_materia, e_assunto, e_banca, e_cargo,
                                  e_ano, e_dif, e_tags, e_questao,
                                  e_gab, e_comentario, e_obs, qid))
                            conn.commit()
                            cursor.close()
                            conn.close()
                            st.success("✅ Questão atualizada com sucesso!")
                            del st.session_state["gq_editando_id"]
                            # Recarrega resultados
                            del st.session_state["gq_resultados"]
                            st.rerun()
                    with col2:
                        if st.button("❌ Cancelar Edição"):
                            del st.session_state["gq_editando_id"]
                            st.rerun()

            else:
                # Exibe a lista de questões
                for q in resultados:
                    qid, materia, assunto, banca, cargo, ano, dif, texto, gab, status, dominio = q
                    preview = (texto or "")[:120] + ("..." if len(texto or "") > 120 else "")

                    with st.container():
                        col_info, col_edit, col_del = st.columns([7, 1, 1])

                        with col_info:
                            st.markdown(
                                f"**#{qid}** — {materia or '—'} | {banca or '—'} "
                                f"| {cargo or '—'} | {ano} | {dif or '—'}  \n"
                                f"*Status:* {status or '—'} | *Domínio:* {dominio or '—'}  \n"
                                f"{preview}"
                            )

                        with col_edit:
                            if st.button("✏️", key=f"edit_{qid}", help="Editar questão"):
                                st.session_state["gq_editando_id"] = qid
                                st.rerun()

                        with col_del:
                            if st.button("🗑️", key=f"del_{qid}", help="Excluir questão"):
                                st.session_state[f"confirm_del_{qid}"] = True
                                st.rerun()

                        # Confirmação de exclusão
                        if st.session_state.get(f"confirm_del_{qid}"):
                            st.warning(f"⚠️ Confirma a exclusão da questão #{qid}?")
                            c1, c2 = st.columns(2)
                            with c1:
                                if st.button("✅ Sim, excluir", key=f"conf_sim_{qid}"):
                                    conn   = get_conn()
                                    cursor = conn.cursor()
                                    cursor.execute("DELETE FROM questoes WHERE id = %s", (qid,))
                                    conn.commit()
                                    cursor.close()
                                    conn.close()
                                    del st.session_state[f"confirm_del_{qid}"]
                                    del st.session_state["gq_resultados"]
                                    st.success(f"Questão #{qid} excluída.")
                                    st.rerun()
                            with c2:
                                if st.button("❌ Cancelar", key=f"conf_nao_{qid}"):
                                    del st.session_state[f"confirm_del_{qid}"]
                                    st.rerun()

                        st.divider()


# ===================================================
# RESOLVER QUESTÕES
# ===================================================

elif menu == "Resolver Questões":

    st.header("Resolver Questões")

    materia_filtro     = st.text_input("Matéria")
    cargo_filtro       = st.text_input("Cargo")
    banca_filtro       = st.text_input("Banca")
    ano_filtro         = st.text_input("Ano")
    dificuldade_filtro = st.selectbox("Dificuldade",
        ["","Muito Fácil","Fácil","Média","Difícil","Muito Difícil","Pegadinha"])
    tags_filtro        = st.multiselect("Tags (pode escolher várias)",
                             options=get_todas_tags())

    col1, col2 = st.columns(2)
    with col1: aleatoria = st.button("🎲 Questão Aleatória")
    with col2: buscar    = st.button("🔍 Buscar Questão")

    if aleatoria:
        conn   = get_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM questoes ORDER BY RANDOM() LIMIT 1")
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()
        if resultado:
            st.session_state["questao"] = resultado
        else:
            st.warning("Nenhuma questão cadastrada ainda.")

    if buscar:
        sql, params = "SELECT * FROM questoes WHERE 1=1", []

        if materia_filtro:
            sql += " AND materia = %s";     params.append(materia_filtro)
        if cargo_filtro:
            sql += " AND cargo = %s";       params.append(cargo_filtro)
        if banca_filtro:
            sql += " AND banca = %s";       params.append(banca_filtro)
        if ano_filtro:
            sql += " AND ano = %s";         params.append(ano_filtro)
        if dificuldade_filtro:
            sql += " AND dificuldade = %s"; params.append(dificuldade_filtro)
        for tag in tags_filtro:
            sql += " AND tags LIKE %s";     params.append(f"%{tag}%")

        sql += " ORDER BY RANDOM() LIMIT 1"

        conn   = get_conn()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()

        if resultado:
            st.session_state["questao"] = resultado
        else:
            st.warning("Nenhuma questão encontrada com esses filtros.")

    if "questao" in st.session_state and st.session_state["questao"]:

        r           = st.session_state["questao"]
        questao_id  = r[0];  materia  = r[1];  assunto    = r[2]
        banca       = r[3];  cargo    = r[4];  ano        = r[5]
        dificuldade = r[6];  tags     = r[7];  questao    = r[8]
        gabarito    = r[9];  comentario = r[10]; observacoes = r[11]
        dominio_atual = r[13] if len(r) > 13 else ""

        st.divider()
        renderizar_com_negrito(questao)

        resposta = st.radio("Sua resposta", ["Certo","Errado"], index=None)

        if st.button("Responder"):

            novo_status = "Acertada" if resposta == gabarito else "Errada"

            if novo_status == "Acertada":
                st.success("✅ Você acertou!")
            else:
                st.error("❌ Você errou!")

            conn   = get_conn()
            cursor = conn.cursor()

            cursor.execute("SELECT status FROM questoes WHERE id = %s", (questao_id,))
            status_atual = (cursor.fetchone() or ["Não respondida"])[0]

            cursor.execute("UPDATE questoes SET status = %s WHERE id = %s",
                           (novo_status, questao_id))

            if status_atual == "Não respondida":
                atualizar_estatisticas_tag(cursor, tags, novo_status)

            conn.commit()
            cursor.close()
            conn.close()

            st.session_state["mostrar_dominio"] = True

        if st.session_state.get("mostrar_dominio"):
            dominio_escolhido = st.radio(
                "Como você avalia seu domínio desta questão?",
                ["🟢 Dominada", "🟡 Revisar", "🔴 Crítica"],
                key="dominio_resolver",
                index=None
            )
            if dominio_escolhido and st.button("Salvar Domínio"):
                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE questoes SET dominio = %s WHERE id = %s",
                               (dominio_escolhido, questao_id))
                conn.commit()
                cursor.close()
                conn.close()
                st.session_state["mostrar_dominio"] = False
                st.success(f"Domínio salvo: {dominio_escolhido}")

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("📖 Ver Comentário"):
                st.session_state["mostrar_comentario"] = True
        with col2:
            if st.button("📝 Minhas Observações"):
                st.session_state["mostrar_obs"] = True
        with col3:
            if st.button("🏷️ Ver Dados da Questão"):
                st.session_state["mostrar_dados"] = True

        if st.session_state.get("mostrar_comentario"):
            st.subheader("Comentário")
            renderizar_com_negrito(comentario)

        if st.session_state.get("mostrar_obs"):
            obs = st.text_area("Minhas Observações", value=observacoes)
            if st.button("Salvar Observação"):
                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE questoes SET observacoes = %s WHERE id = %s",
                               (obs, questao_id))
                conn.commit()
                cursor.close()
                conn.close()
                st.success("Observação salva.")

        if st.session_state.get("mostrar_dados"):
            st.subheader("Dados da Questão")
            st.write(f"**Banca:** {banca}")
            st.write(f"**Cargo:** {cargo}")
            st.write(f"**Ano:** {ano}")
            st.write(f"**Dificuldade:** {dificuldade}")
            st.write("**Tags:**")
            for tag in (tags or "").splitlines():
                if tag.strip(): st.write(f"• {tag}")


# ===================================================
# REVISAR QUESTÕES
# ===================================================

elif menu == "Revisar Questões":

    st.header("Revisar Questões")

    materia_filtro     = st.text_input("Matéria",    key="rev_materia")
    cargo_filtro       = st.text_input("Cargo",      key="rev_cargo")
    banca_filtro       = st.text_input("Banca",      key="rev_banca")
    ano_filtro         = st.text_input("Ano",        key="rev_ano")
    dificuldade_filtro = st.selectbox("Dificuldade",
        ["","Muito Fácil","Fácil","Média","Difícil","Muito Difícil","Pegadinha"],
        key="rev_dificuldade")
    status_filtro      = st.selectbox("Status",
        ["","Não respondida","Acertada","Errada"], key="rev_status")

    dominio_filtro = st.selectbox("Domínio",
        ["","🟢 Dominada","🟡 Revisar","🔴 Crítica"], key="rev_dominio")

    todas_tags   = get_todas_tags()
    default_tags = []
    if "tag_revisao" in st.session_state:
        tag_pre = st.session_state.pop("tag_revisao")
        if tag_pre in todas_tags:
            default_tags = [tag_pre]

    tags_filtro = st.multiselect("Tags (pode escolher várias)",
                      options=todas_tags, default=default_tags, key="rev_tags")

    if st.button("🔍 Buscar Questões"):

        sql, params = "SELECT * FROM questoes WHERE 1=1", []

        if materia_filtro:
            sql += " AND materia = %s";     params.append(materia_filtro)
        if cargo_filtro:
            sql += " AND cargo = %s";       params.append(cargo_filtro)
        if banca_filtro:
            sql += " AND banca = %s";       params.append(banca_filtro)
        if ano_filtro:
            sql += " AND ano = %s";         params.append(ano_filtro)
        if dificuldade_filtro:
            sql += " AND dificuldade = %s"; params.append(dificuldade_filtro)
        if status_filtro:
            sql += " AND status = %s";      params.append(status_filtro)
        if dominio_filtro:
            sql += " AND dominio = %s";     params.append(dominio_filtro)
        for tag in tags_filtro:
            sql += " AND tags LIKE %s";     params.append(f"%{tag}%")

        conn   = get_conn()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        questoes_filtradas = cursor.fetchall()
        cursor.close()
        conn.close()

        st.session_state["questoes_revisao"] = questoes_filtradas
        st.session_state["indice_revisao"]   = 0

        if not questoes_filtradas:
            st.warning("Nenhuma questão encontrada com esses filtros.")

    if (
        "questoes_revisao" in st.session_state
        and len(st.session_state["questoes_revisao"]) > 0
    ):
        total  = len(st.session_state["questoes_revisao"])
        indice = st.session_state["indice_revisao"]

        st.success(f"{total} questões encontradas.")

        q           = st.session_state["questoes_revisao"][indice]
        questao_id  = q[0];  materia  = q[1];  assunto    = q[2]
        banca       = q[3];  cargo    = q[4];  ano        = q[5]
        dificuldade = q[6];  tags     = q[7];  texto      = q[8]
        gabarito    = q[9];  comentario = q[10]; observacoes = q[11]
        dominio_atual = q[13] if len(q) > 13 else ""

        st.divider()
        renderizar_com_negrito(texto)

        resposta = st.radio("Sua resposta", ["Certo","Errado"],
                            index=None, key=f"resp_{questao_id}")

        if st.button("Responder Revisão"):

            novo_status = "Acertada" if resposta == gabarito else "Errada"

            if novo_status == "Acertada":
                st.success("✅ Você acertou!")
            else:
                st.error("❌ Você errou!")

            conn   = get_conn()
            cursor = conn.cursor()

            cursor.execute("SELECT status FROM questoes WHERE id = %s", (questao_id,))
            status_atual = (cursor.fetchone() or ["Não respondida"])[0]

            cursor.execute("UPDATE questoes SET status = %s WHERE id = %s",
                           (novo_status, questao_id))

            if status_atual == "Não respondida":
                atualizar_estatisticas_tag(cursor, tags, novo_status)

            conn.commit()
            cursor.close()
            conn.close()

            st.session_state["mostrar_dominio_rev"] = True

        if st.session_state.get("mostrar_dominio_rev"):
            idx_dominio = None
            opcoes_dominio = ["🟢 Dominada", "🟡 Revisar", "🔴 Crítica"]
            if dominio_atual in opcoes_dominio:
                idx_dominio = opcoes_dominio.index(dominio_atual)
            dominio_escolhido = st.radio(
                "Como você avalia seu domínio desta questão?",
                opcoes_dominio,
                key=f"dominio_rev_{questao_id}",
                index=idx_dominio
            )
            if dominio_escolhido and st.button("Salvar Domínio", key=f"salvar_dom_{questao_id}"):
                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE questoes SET dominio = %s WHERE id = %s",
                               (dominio_escolhido, questao_id))
                conn.commit()
                cursor.close()
                conn.close()
                st.session_state["mostrar_dominio_rev"] = False
                st.success(f"Domínio salvo: {dominio_escolhido}")

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("📖 Comentário", key=f"coment_{questao_id}"):
                st.session_state["mostrar_comentario_rev"] = True
        with col2:
            if st.button("📝 Observações", key=f"obs_{questao_id}"):
                st.session_state["mostrar_obs_rev"] = True
        with col3:
            if st.button("🏷️ Dados", key=f"dados_{questao_id}"):
                st.session_state["mostrar_dados_rev"] = True

        if st.session_state.get("mostrar_comentario_rev"):
            st.subheader("Comentário")
            renderizar_com_negrito(comentario)

        if st.session_state.get("mostrar_obs_rev"):
            obs = st.text_area("Minhas Observações", value=observacoes,
                               key=f"obs_text_{questao_id}")
            if st.button("Salvar Observação", key=f"salvar_obs_{questao_id}"):
                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE questoes SET observacoes = %s WHERE id = %s",
                               (obs, questao_id))
                conn.commit()
                cursor.close()
                conn.close()
                st.success("Observação salva.")

        if st.session_state.get("mostrar_dados_rev"):
            st.subheader("Dados da Questão")
            st.write(f"**Banca:** {banca}")
            st.write(f"**Cargo:** {cargo}")
            st.write(f"**Ano:** {ano}")
            st.write(f"**Dificuldade:** {dificuldade}")
            st.write("**Tags:**")
            for tag in (tags or "").splitlines():
                if tag.strip(): st.write(f"• {tag}")

        col1, col2 = st.columns(2)
        with col1:
            if indice > 0 and st.button("⬅️ Anterior"):
                st.session_state["indice_revisao"] -= 1
                st.rerun()
        with col2:
            if indice < total - 1 and st.button("➡️ Próxima"):
                st.session_state["indice_revisao"] += 1
                st.rerun()


# ===================================================
# ESTATÍSTICAS
# ===================================================

elif menu == "Estatísticas":

    st.header("Estatísticas por Tag")

    lista_tags = get_todas_tags()

    if not lista_tags:
        st.warning("Nenhuma tag cadastrada.")

    else:
        tag_escolhida = st.selectbox("Escolha uma Tag", lista_tags)

        conn   = get_conn()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT ciclo FROM estatisticas_tag
            WHERE tag = %s ORDER BY ciclo
        """, (tag_escolhida,))
        ciclos = [r[0] for r in cursor.fetchall()]

        if not ciclos:
            ciclos = [1]
            cursor.execute("""
                INSERT INTO estatisticas_tag (tag, ciclo, acertos, erros)
                VALUES (%s, 1, 0, 0)
            """, (tag_escolhida,))
            conn.commit()

        ciclo_escolhido = st.selectbox("Ciclo", ciclos)

        cursor.execute("""
            SELECT acertos, erros FROM estatisticas_tag
            WHERE tag = %s AND ciclo = %s
        """, (tag_escolhida, ciclo_escolhido))
        res     = cursor.fetchone()
        acertos = res[0] if res else 0
        erros   = res[1] if res else 0
        total   = acertos + erros
        pct     = (acertos / total * 100) if total > 0 else 0

        cursor.execute("SELECT COUNT(*) FROM questoes WHERE tags LIKE %s",
                       (f"%{tag_escolhida}%",))
        total_questoes_tag = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        st.divider()
        st.subheader(f"🏷️ {tag_escolhida}")
        st.write(f"🔥 Nível de domínio: {nivel_dominio(pct)}")
        st.write(f"📚 Ciclo: {ciclo_escolhido}")

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total da Tag",      total_questoes_tag)
        c2.metric("Questões do Ciclo", total)
        c3.metric("Acertos",           acertos)
        c4.metric("Erros",             erros)
        c5.metric("% Acerto",          f"{pct:.2f}%")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("📚 Revisar Somente Esta Tag"):
                st.session_state["tag_revisao"]   = tag_escolhida
                st.session_state["menu_override"] = "Revisar Questões"
                st.session_state["rev_status"]    = ""
                st.rerun()

        with col2:
            if st.button("🔄 Criar Novo Ciclo"):
                novo_ciclo = max(ciclos) + 1
                conn   = get_conn()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO estatisticas_tag (tag, ciclo, acertos, erros)
                    VALUES (%s, %s, 0, 0)
                """, (tag_escolhida, novo_ciclo))
                cursor.execute("""
                    UPDATE questoes SET status = 'Não respondida'
                    WHERE tags LIKE %s
                """, (f"%{tag_escolhida}%",))
                conn.commit()
                cursor.close()
                conn.close()
                st.success(f"Ciclo {novo_ciclo} criado.")
                st.rerun()


# ===================================================
# DASHBOARD
# ===================================================

elif menu == "Dashboard":

    st.header("📊 Dashboard das Tags")

    lista_tags = get_todas_tags()

    if not lista_tags:
        st.warning("Nenhuma tag cadastrada ainda.")

    else:
        conn   = get_conn()
        cursor = conn.cursor()

        dados = []
        for tag in lista_tags:
            cursor.execute("""
                SELECT ciclo, acertos, erros FROM estatisticas_tag
                WHERE tag = %s ORDER BY ciclo DESC LIMIT 1
            """, (tag,))
            linha = cursor.fetchone()
            ciclo, acertos, erros = linha if linha else (1, 0, 0)
            total = acertos + erros
            pct   = (acertos / total * 100) if total > 0 else 0

            cursor.execute("SELECT COUNT(*) FROM questoes WHERE tags LIKE %s",
                           (f"%{tag}%",))
            total_banco = cursor.fetchone()[0]

            dados.append({
                "tag": tag, "ciclo": ciclo, "acertos": acertos,
                "erros": erros, "total_ciclo": total,
                "pct": pct, "total_banco": total_banco,
                "cor": cor_dominio(pct),
            })

        cursor.close()
        conn.close()

        tags_labels = [d["tag"] for d in dados]
        pcts        = [d["pct"] for d in dados]

        st.subheader("% de Acerto por Tag (ciclo mais recente)")
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=tags_labels, y=pcts,
            marker_color=[d["cor"] for d in dados],
            text=[f"{p:.1f}%" for p in pcts],
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>% Acerto: %{y:.1f}%<extra></extra>"
        ))
        fig.add_hline(y=70, line_dash="dot", line_color="orange",
                      annotation_text="Meta mínima (70%)",
                      annotation_position="bottom right")
        fig.update_layout(yaxis=dict(title="% Acerto", range=[0, 110]),
                          xaxis=dict(title="Tag"), showlegend=False,
                          height=420, margin=dict(t=40, b=60))
        st.plotly_chart(fig, use_container_width=True)

        st.subheader("Acertos vs Erros por Tag")
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(name="Acertos", x=tags_labels,
                              y=[d["acertos"] for d in dados],
                              marker_color="#2ecc71"))
        fig2.add_trace(go.Bar(name="Erros", x=tags_labels,
                              y=[d["erros"] for d in dados],
                              marker_color="#e74c3c"))
        fig2.update_layout(barmode="stack",
                           yaxis=dict(title="Questões"), xaxis=dict(title="Tag"),
                           height=380, margin=dict(t=40, b=60))
        st.plotly_chart(fig2, use_container_width=True)

        st.subheader("Resumo por Tag")
        st.markdown("| Tag | Ciclo | Total no Banco | Respondidas | Acertos | Erros | % Acerto | Domínio |")
        st.markdown("|-----|-------|---------------|-------------|---------|-------|----------|---------|")
        for d in dados:
            st.markdown(
                f"| {d['tag']} | {d['ciclo']} | {d['total_banco']} "
                f"| {d['total_ciclo']} | {d['acertos']} | {d['erros']} "
                f"| {d['pct']:.1f}% | {nivel_dominio(d['pct'])} |"
            )
